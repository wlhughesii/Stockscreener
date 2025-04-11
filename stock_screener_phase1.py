import os
import datetime
import pandas as pd
import yfinance as yf
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tradier_config import get_api_key, get_base_url
from fmp_config import get_fmp_key

# === CONFIG ===
CSP_SCORE_THRESHOLD = 6.0
NEAR_MISS_LOWER_BOUND = 5.5

def get_fmp_fundamentals(ticker, api_key):
    try:
        url = f"https://financialmodelingprep.com/api/v3/key-metrics-ttm/{ticker}?apikey={api_key}"
        response = requests.get(url)
        data = response.json()
        if isinstance(data, list) and len(data) > 0:
            metrics = data[0]
            return {
                "pe": metrics.get("peRatioTTM"),
                "net_margin": metrics.get("netProfitMarginTTM"),
                "roe": metrics.get("roeTTM")
            }
    except Exception as e:
        print(f"[{ticker}] FMP error: {e}")
    return {"pe": None, "net_margin": None, "roe": None}

def fallback_yf_fundamentals(ticker):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        return {
            "pe": info.get("trailingPE"),
            "net_margin": info.get("netMargins"),
            "roe": info.get("returnOnEquity")
        }
    except Exception as e:
        print(f"[{ticker}] YF fallback error: {e}")
    return {"pe": None, "net_margin": None, "roe": None}
def is_etf_from_profile(ticker):
    try:
        url = f"https://financialmodelingprep.com/api/v3/profile/{ticker}?apikey={get_fmp_key()}"
        response = requests.get(url)
        data = response.json()
        if isinstance(data, list) and len(data) > 0:
            return data[0].get("isEtf", False)
    except Exception as e:
        print(f"[{ticker}] ETF check failed: {e}")
    return False

def calculate_score(pe, net_margin, roe):
    score = 0
    explanation = []
    metric_count = 0

    # PE Ratio
    if pe is not None:
        metric_count += 1
        if pe < 20:
            score += 3
            explanation.append(f"PE = {pe:.2f} → Strong")
        elif pe < 30:
            score += 2
            explanation.append(f"PE = {pe:.2f} → Decent")
        else:
            score += 1
            explanation.append(f"PE = {pe:.2f} → High")

    # Net Margin
    if net_margin is not None:
        metric_count += 1
        if net_margin > 0.15:
            score += 3
            explanation.append(f"Net Margin = {net_margin:.2%} → Excellent")
        elif net_margin > 0.1:
            score += 2
            explanation.append(f"Net Margin = {net_margin:.2%} → Good")
        else:
            score += 1
            explanation.append(f"Net Margin = {net_margin:.2%} → Weak")

    # ROE
    if roe is not None:
        metric_count += 1
        if roe > 0.2:
            score += 3
            explanation.append(f"ROE = {roe:.2%} → Excellent")
        elif roe > 0.1:
            score += 2
            explanation.append(f"ROE = {roe:.2%} → Good")
        else:
            score += 1
            explanation.append(f"ROE = {roe:.2%} → Weak")

    # Normalize score if fewer metrics available
    if metric_count > 0:
        score = round(score * (3 / (metric_count * 3)) * 9, 2)
    else:
        score = 0
        explanation.append("No fundamentals available")

    return score, "; ".join(explanation)

def color_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        score_cell = row[1]  # Score
        try:
            score = float(score_cell.value)
            if score >= 8:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            elif score >= CSP_SCORE_THRESHOLD:
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
            else:
                fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")  # Red
            for cell in row:
                cell.fill = fill
        except:
            continue
    wb.save(filepath)
def main():
    print("📊 Running Enhanced Stock Screener Phase 1...")
    tickers_file = "tickers.txt"
    if not os.path.exists(tickers_file):
        print("❌ Missing tickers.txt file.")
        return

    with open(tickers_file, "r") as f:
        tickers = [line.strip().upper() for line in f if line.strip()]

    api_key = get_api_key()
    base_url = get_base_url()
    fmp_key = get_fmp_key()

    rows = []
    for ticker in tickers:
        print(f"→ Analyzing {ticker}...")
        is_etf_flag = is_etf_from_profile(ticker)

        fundamentals = get_fmp_fundamentals(ticker, fmp_key)
        if not any(fundamentals.values()):
            fundamentals = fallback_yf_fundamentals(ticker)

        pe = fundamentals.get("pe")
        net_margin = fundamentals.get("net_margin")
        roe = fundamentals.get("roe")

        score, explanation = calculate_score(pe, net_margin, roe)

        # Get price from Tradier
        price = 0
        try:
            headers = {"Authorization": f"Bearer {api_key}", "Accept": "application/json"}
            url = f"{base_url}markets/quotes?symbols={ticker}"
            r = requests.get(url, headers=headers)
            data = r.json()
            quote = data["quotes"]["quote"]
            price = quote.get("last", 0)
        except Exception as e:
            print(f"[{ticker}] Quote error: {e}")

        rows.append({
            "Ticker": ticker,
            "Score": score,
            "Price": round(price, 2),
            "PE Ratio": round(pe, 2) if pe else "N/A",
            "Net Margin": round(net_margin, 4) if net_margin else "N/A",
            "ROE": round(roe, 4) if roe else "N/A",
            "ETF": "Yes" if is_etf_flag else "No",
            "Scoring Explanation": explanation,
            "CSP Qualified": "Yes" if score >= CSP_SCORE_THRESHOLD else "No",
            "Near Miss": "Yes" if NEAR_MISS_LOWER_BOUND <= score < CSP_SCORE_THRESHOLD else "No"
        })

    df = pd.DataFrame(rows)
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    os.makedirs("results", exist_ok=True)
    filename = f"results/CSP_Phase1_Top50_{now}.xlsx"
    df.to_excel(filename, index=False)
    color_excel(filename)
    print(f"✅ Screener complete. Results saved to: {filename}")

if __name__ == "__main__":
    main()
